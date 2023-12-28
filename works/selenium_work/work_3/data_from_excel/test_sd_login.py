from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from constants import globalConstants as gc
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl
import pytest


#Bir önceki ödevde yapılan işlemler excel dosyasından veri alınarak yazılmıştır. (work_2)

class test_sauceDemo:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.get(gc.BASE_URL)
        self.driver.maximize_window() 
    
    
    def getInvalidData():
        invalid_excel = openpyxl.load_workbook("data/login_test.xlsx")
        sheet = invalid_excel["invalid_login"]
        rows = sheet.max_row
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data
    
    def getValidData():
        valid_excel = openpyxl.load_workbook("data/login_test.xlsx")
        sheet = valid_excel["valid_login"]
        rows = sheet.max_row
        data = []
        for i in range(2, rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data  
    
    def get_Pasword_EmptyData():
        passempty_excel = openpyxl.load_workbook("data/login_test.xlsx")
        sheet = passempty_excel["password_empty"]
        rows = sheet.max_row
        data = []
        for i in range(2, rows+1):
            username = sheet.cell(i,1).value
            data.append((username))
        return data  
    
    def getErrorData():
        error_excel = openpyxl.load_workbook("data/login_test.xlsx")
        sheet = error_excel["error"]
        rows = sheet.max_row
        data = []
        for i in range(2, rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data  
    
    def getCheckout():
        checkout_excel = openpyxl.load_workbook("data/login_test.xlsx")
        sheet = checkout_excel["checkout"]
        rows = sheet.max_row
        data = []
        for i in range(4, rows+1):
            firstname = sheet.cell(i,1).value
            lastname = sheet.cell(i,2).value
            postalcode = sheet.cell(i,3).value
            data.append((firstname,lastname,postalcode))
        return data  
    
    #********************************#
    
    @pytest.mark.parametrize("username, password" , getInvalidData())
    def test_username_password_empty(self,username,password):
        self.setup_method()
        usernameInput =  WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.USERNAME_ID)))
        usernameInput.send_keys(username)
        sleep(2)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.PASSWORD_ID)))
        passwordInput.send_keys(password)
        sleep(2)
        loginButton = self.find_element(By.ID , gc.LOGIN_BUTTON)
        loginButton.click()
        sleep(3)
        errorMessage = self.find_element(By.XPATH, gc.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == gc.EMPTY_INVALID
    
     #********************************#
    
    @pytest.mark.parametrize("username,password" , get_Pasword_EmptyData())
    def test_password_empty(self,username,password):
        self.setup_method()
        usernameInput =  WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.USERNAME_ID)))
        usernameInput.send_keys(username)
        sleep(2)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.PASSWORD_ID)))
        passwordInput.send_keys(password)
        sleep(2)
        loginButton = self.find_element(By.ID , gc.LOGIN_BUTTON)
        loginButton.click()
        sleep(3)
        errorMessage = self.find_element(By.XPATH, gc.ERROR_MESSAGE_XPATH)
        assert  errorMessage.text == "Epic sadface: Password is required"
    
     #********************************#
    
    @pytest.mark.parametrize("username,password" , getErrorData())
    def test_error(self,username,password):
        self.setup_method()
        usernameInput =  WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.USERNAME_ID)))
        usernameInput.send_keys(username)
        sleep(2)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.PASSWORD_ID)))
        passwordInput.send_keys(password)
        sleep(2)
        loginButton = self.find_element(By.ID , gc.LOGIN_BUTTON)
        loginButton.click()
        sleep(3)
        errorMessage = self.find_element(By.XPATH, gc.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == "Epic sadface: Sorry, this user has been locked out."
    
     #********************************#
    
    @pytest.mark.parametrize("username, password" , getValidData())
    def test_successfull_login(self,username,password):
        self.setup_method()
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.USERNAME_ID)))
        usernameInput.send_keys(username)
        sleep(2)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.PASSWORD_ID)))
        passwordInput.send_keys(password)
        sleep(2)
        loginButton = self.find_element(By.ID , gc.LOGIN_BUTTON)
        loginButton.click()
        sleep(3)
        listOfElement = self.find_elements(By.CLASS_NAME, gc.INVENTORY_ITEM)
        assert len(listOfElement)==6
    
     #********************************#
        
    @pytest.mark.parametrize("firstname,lastname,postalcode" , )
    def test_checkout(self,firstname,lastname,postalcode):
        self.setup_method()
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        sleep(2)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        sleep(2)
        loginButton = self.find_element(By.ID , gc.LOGIN_BUTTON)
        loginButton.click()
        shopping_list = self.driver.find_element(By.XPATH, gc.SHOPPIND_XPATH)
        shopping_list.click()
        sleep(2)
        checkout = self.driver.find_element(By.ID, gc.CHECKOUT)
        checkout.click()
        sleep(2)
        first_name = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID , gc.FIRST_NAME)))
        first_name.send_keys(firstname)
        last_name = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID , gc.LAST_NAME)))
        last_name.send_keys(lastname)
        postal_code = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID , gc.POSTAL_CODE)))
        postal_code.send_keys(postalcode)
        button_continue = self.driver.find_element(By.ID ,gc.CONTINUE )
        button_continue.click()
        errorMessage = self.driver.find_element(By.XPATH, gc.CHECKOUT_ERROR_XPATH)
        assert errorMessage.text == "Error: First Name is required"