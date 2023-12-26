import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager . chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #beklediğim koşulu sağlayana kadar bekle, ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from constants import globalConstants as gc

class test_DemoClass2:
    def setup_method(self): #her test başlangıcında çalışacak fonksiyon.
        self.driver = webdriver.Chrome()
        self.driver.get(gc.BASE_URL)
        self.driver.maximize_window() 
                
    
    def getData():
        
        #EXCEL KULLANIMI
        excel = openpyxl.load_workbook("data/invalid_login.xlsx")
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağımı göstermeme yardımcı olur.
        rows = sheet.max_row #kaçıncı satıra kadar veri var ? 
        data = []
        for i in range(2,rows):
            username = sheet.cell(i,1).value #hücreni satırı ve sutununu verir. Burada i=rows , 1 = A sutunu
            password = sheet.cell(i,2).value #hücreni satırı ve sutununu verir. Burada i=rows , 2 = B sutunu
            data.append((username,password)) #çift parantaz username ve passwordu bir data olarak alır.
        
        return data
    
    
    @pytest.mark.parametrize("username,password" , getData())
    def test_invalid_login(self, username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID , gc.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID , gc.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID , gc.LOGIN_BUTTON)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH, gc.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == gc.ERROR_MESSAGE_XPATH
        
    @pytest.mark.skip  
    def test_valid_login(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID , gc.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID , gc.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID , gc.LOGIN_BUTTON)
        loginButton.click()
        headerLogo = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH , "//*[@id='header_container']/div[1]/div[2]/div")))
        testResult = headerLogo.text == "Swag Labs"
        print(f"Test Sonucu: {testResult}")

        
    #testleri yaparken test ön kısmını eklememiz gerekir. Yoksa sistemimiz test edilemez.
    #prefix => test_ 